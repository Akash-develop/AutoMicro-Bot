import json
import os
from langchain_core.tools import tool
from app.graph.tools.permission_manager import check_permission

@tool
@check_permission("create_excel_with_sample_data")
def create_excel_with_sample_data(file_path: str, data_json: str = "") -> str:
    """
    Creates a real Excel (.xlsx) file.

    - file_path: where to save the file (e.g. '/Users/me/report.xlsx').
      A .xlsx extension is added automatically if missing.
    - data_json: optional JSON string describing the content.
      Format: {"headers": ["Col1", "Col2"], "rows": [[val1, val2], ...]}
      If omitted, a sample employee table is created.

    Bug #10 fix: previously created a CSV with hardcoded data.
    Now creates a genuine Excel file using openpyxl and accepts
    user-specified data via the data_json parameter.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Ensure .xlsx extension
        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"

        # ── Parse data_json or fall back to sample data ──────────────────────
        headers = ["ID", "Name", "Role", "Department"]
        rows = [
            [1, "Alice Smith",    "Software Engineer", "Engineering"],
            [2, "Bob Johnson",    "Product Manager",   "Product"],
            [3, "Charlie Brown",  "UX Designer",       "Design"],
            [4, "Diana Prince",   "Data Scientist",     "Data"],
        ]

        if data_json and data_json.strip():
            try:
                parsed = json.loads(data_json)
                if "headers" in parsed:
                    headers = parsed["headers"]
                if "rows" in parsed:
                    rows = parsed["rows"]
            except json.JSONDecodeError as je:
                return f"Invalid data_json format (must be valid JSON): {je}"

        # ── Build workbook ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header row — bold, light blue background
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths (approximate)
        for col_idx, header in enumerate(headers, start=1):
            col_values = [str(header)] + [str(r[col_idx - 1]) for r in rows if col_idx - 1 < len(r)]
            best_width = max(len(v) for v in col_values) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = best_width

        # Ensure the directory exists
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)

        wb.save(file_path)
        return (
            f"Excel file saved to '{file_path}' "
            f"({len(rows)} data rows, {len(headers)} columns)."
        )

    except Exception as e:
        return f"Failed to create Excel file: {str(e)}"
